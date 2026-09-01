from typing import List
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta SAIP ────────────────────────────────────────────────────────────
COLOR_HEADER_BG = "003459"
COLOR_HEADER_FG = "FFFFFF"
COLOR_ROW_ALT = "D9DCD6"
COLOR_ROW_WHITE = "FFFFFF"
COLOR_BORDER = "CFC0B0"


def header_font():
    return Font(name="Arial", bold=True, color=COLOR_HEADER_FG, size=10)


def cell_font():
    return Font(name="Arial", size=9)


def header_fill():
    return PatternFill("solid", fgColor=COLOR_HEADER_BG)


def alt_fill():
    return PatternFill("solid", fgColor=COLOR_ROW_ALT)


def white_fill():
    return PatternFill("solid", fgColor=COLOR_ROW_WHITE)


def border():
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def center():
    return Alignment(horizontal="center", vertical="center")


def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def apply_header_row(ws, headers: List[str]):
    if ws is None:
        return
    ws.row_dimensions[1].height = 22
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.border = border()
        cell.alignment = center()


def apply_data_row(ws, row_idx: int, values: list):
    if ws is None:
        return
    fill = alt_fill() if row_idx % 2 == 0 else white_fill()
    ws.row_dimensions[row_idx].height = 18
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font()
        cell.fill = fill
        cell.border = border()
        cell.alignment = left()


def auto_width(ws, headers: List[str]):
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header) + 4
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[col_letter].width = min(max_len, 40)


def add_title_row(ws, title: str, col_count: int):
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(name="Arial", bold=True, size=12, color=COLOR_HEADER_FG)
    cell.fill = header_fill()
    cell.alignment = center()
    ws.row_dimensions[1].height = 28


def fmt_dt(value) -> str:
    if value is None:
        return "—"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value)


def fmt_bool(value: bool) -> str:
    return "Sí" if value else "No"
