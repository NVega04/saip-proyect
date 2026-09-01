from typing import Optional, List, Literal
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from io import BytesIO
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.database import get_session
from src.dependencies import require_module
from src.models.models import (
    Product,
    Supply,
    CommercialProduct,
    User,
)
from src.schemas.schemas import (
    InventoryItem,
    InventoryItemStatus,
    InventoryPage,
    InventorySummary,
)
from src.bulk.parser import (
    template_response,
)

router = APIRouter(prefix="/inventory", tags=["Inventory"])

# ── Paleta SAIP ────────────────────────────────────────────────────────────
COLOR_HEADER_BG = "5C3D1E"
COLOR_HEADER_FG = "FFFFFF"
COLOR_ROW_ALT   = "F5F0EA"
COLOR_ROW_WHITE = "FFFFFF"
COLOR_BORDER    = "CFC0B0"


def _header_font():
    return Font(name="Arial", bold=True, color=COLOR_HEADER_FG, size=10)

def _cell_font():
    return Font(name="Arial", size=9)

def _header_fill():
    return PatternFill("solid", fgColor=COLOR_HEADER_BG)

def _alt_fill():
    return PatternFill("solid", fgColor=COLOR_ROW_ALT)

def _white_fill():
    return PatternFill("solid", fgColor=COLOR_ROW_WHITE)

def _border():
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)

def _center():
    return Alignment(horizontal="center", vertical="center")

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _apply_header_row(ws, headers: List[str]):
    ws.row_dimensions[1].height = 22
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = _header_font()
        cell.fill      = _header_fill()
        cell.border    = _border()
        cell.alignment = _center()


def _apply_data_row(ws, row_idx: int, values: list):
    fill = _alt_fill() if row_idx % 2 == 0 else _white_fill()
    ws.row_dimensions[row_idx].height = 18
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font      = _cell_font()
        cell.fill      = fill
        cell.border    = _border()
        cell.alignment = _left()


def _auto_width(ws, headers: List[str]):
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header) + 4
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[col_letter].width = min(max_len, 40)


def _add_title_row(ws, title: str, col_count: int):
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font      = Font(name="Arial", bold=True, size=12, color=COLOR_HEADER_FG)
    cell.fill      = _header_fill()
    cell.alignment = _center()
    ws.row_dimensions[1].height = 28


# ── Consulta consolidada ───────────────────────────────────────────────────
def _status_of(available: float, min_stock: float, max_stock: float) -> InventoryItemStatus:
    if available <= min_stock:
        return InventoryItemStatus.LOW
    if max_stock > 0 and available > max_stock:
        return InventoryItemStatus.OVER
    return InventoryItemStatus.NORMAL


def _build_rows(
    session: Session,
    item_type: Optional[str],
    stock_status: Optional[str],
    category_id: Optional[int],
    search: Optional[str],
    status_filter: Optional[str],
) -> List[InventoryItem]:
    rows: List[InventoryItem] = []

    def _matches(available: float, min_stock: float, max_stock: float) -> bool:
        if stock_status is None:
            return True
        return _status_of(available, min_stock, max_stock).value == stock_status

    if item_type in (None, "product"):
        q = select(Product).where(Product.deleted_at == None)
        if status_filter is not None:
            q = q.where(Product.status == status_filter)
        if search:
            q = q.where(Product.name.contains(search))
        for p in session.exec(q).all():
            if not _matches(p.available_quantity, p.min_stock, p.max_stock):
                continue
            rows.append(InventoryItem(
                id=p.id,
                item_type="product",
                name=p.name,
                description=p.description,
                category_name=None,
                unit_abbreviation=p.unit.abbreviation if p.unit else None,
                available_quantity=p.available_quantity,
                min_stock=p.min_stock,
                max_stock=p.max_stock,
                stock_status=_status_of(p.available_quantity, p.min_stock, p.max_stock),
                status=p.status,
            ))

    if item_type in (None, "supply"):
        q = select(Supply).where(Supply.deleted_at == None)
        if status_filter is not None:
            q = q.where(Supply.status == status_filter)
        if category_id is not None:
            q = q.where(Supply.category_id == category_id)
        if search:
            q = q.where(Supply.name.contains(search))
        for s in session.exec(q).all():
            if not _matches(s.available_quantity, s.min_stock, s.max_stock):
                continue
            rows.append(InventoryItem(
                id=s.id,
                item_type="supply",
                name=s.name,
                description=s.description,
                category_name=s.category.name if s.category else None,
                unit_abbreviation=s.unit.abbreviation if s.unit else None,
                available_quantity=s.available_quantity,
                min_stock=s.min_stock,
                max_stock=s.max_stock,
                stock_status=_status_of(s.available_quantity, s.min_stock, s.max_stock),
                status=s.status,
            ))

    if item_type in (None, "commercial"):
        q = select(CommercialProduct).where(CommercialProduct.deleted_at == None)
        if status_filter is not None:
            q = q.where(CommercialProduct.status == status_filter)
        if category_id is not None:
            q = q.where(CommercialProduct.category_id == category_id)
        if search:
            q = q.where(CommercialProduct.name.contains(search))
        for c in session.exec(q).all():
            if not _matches(c.available_quantity, c.min_stock, c.max_stock):
                continue
            rows.append(InventoryItem(
                id=c.id,
                item_type="commercial",
                name=c.name,
                description=c.description,
                category_name=c.category.name if c.category else None,
                unit_abbreviation=c.unit.abbreviation if c.unit else None,
                available_quantity=c.available_quantity,
                min_stock=c.min_stock,
                max_stock=c.max_stock,
                stock_status=_status_of(c.available_quantity, c.min_stock, c.max_stock),
                status=c.status,
            ))

    rows.sort(key=lambda r: r.name.lower())
    return rows


# ── Endpoints ──────────────────────────────────────────────────────────────
@router.get("/summary", response_model=InventoryPage, status_code=status.HTTP_200_OK)
def inventory_summary(
    item_type: Optional[Literal["supply", "product", "commercial"]] = Query(default=None),
    stock_status: Optional[Literal["bajo", "normal", "sobre"]] = Query(default=None),
    category_id: Optional[int] = Query(default=None),
    search: Optional[str] = Query(default=None, max_length=150),
    status_filter: Optional[Literal["active", "inactive"]] = Query(default=None, alias="status"),
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=10, ge=1, le=100),
    session: Session = Depends(get_session),
    current_user: User = Depends(require_module("inventory")),
):
    if category_id is not None and item_type == "product":
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="El filtro de categoría solo aplica a insumos y productos comerciales.",
        )

    rows = _build_rows(session, item_type, stock_status, category_id, search, status_filter)
    total = len(rows)

    summary = InventorySummary(
        total_items=total,
        bajo_stock=sum(1 for r in rows if r.stock_status == InventoryItemStatus.LOW),
        sobre_stock=sum(1 for r in rows if r.stock_status == InventoryItemStatus.OVER),
    )

    start = (page - 1) * limit
    return InventoryPage(
        items=rows[start : start + limit],
        total=total,
        page=page,
        limit=limit,
        summary=summary,
    )


@router.get("/report", response_class=StreamingResponse)
def inventory_report(
    item_type: Optional[Literal["supply", "product", "commercial"]] = Query(default=None),
    stock_status: Optional[Literal["bajo", "normal", "sobre"]] = Query(default=None),
    category_id: Optional[int] = Query(default=None),
    search: Optional[str] = Query(default=None, max_length=150),
    status_filter: Optional[Literal["active", "inactive"]] = Query(default=None, alias="status"),
    session: Session = Depends(get_session),
    current_user: User = Depends(require_module("inventory")),
):
    if category_id is not None and item_type == "product":
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="El filtro de categoría solo aplica a insumos y productos comerciales.",
        )

    rows = _build_rows(session, item_type, stock_status, category_id, search, status_filter)

    headers = ["Tipo", "Nombre", "Descripción", "Categoría", "Unidad",
               "Stock disponible", "Stock mínimo", "Stock máximo",
               "Estado de stock", "Estado"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.freeze_panes = "A3"

    _apply_header_row(ws, headers)
    for i, r in enumerate(rows, start=2):
        _apply_data_row(ws, i, [
            {"supply": "Insumo", "product": "Producto terminado", "commercial": "Comercial"}[r.item_type],
            r.name,
            r.description or "—",
            r.category_name or "—",
            r.unit_abbreviation or "—",
            r.available_quantity,
            r.min_stock,
            r.max_stock,
            {"bajo": "Bajo stock", "normal": "Normal", "sobre": "Sobre stock"}[r.stock_status.value],
            r.status,
        ])

    _add_title_row(
        ws,
        f"Reporte de Inventario — {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC",
        len(headers),
    )
    _auto_width(ws, headers)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"reporte_inventario_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Carga masiva de stock (entradas/descuentos) ─────────────────────────────
BULK_MOVEMENT_HEADERS = ["tipo", "nombre", "cantidad", "nota"]
BULK_MOVEMENT_EXAMPLES = [
    ["Insumo", "Harina de trigo", 50, "Compra semanal"],
    ["Producto", "Pan frances", 10, "Produccion extra"],
]


@router.get(
    "/bulk/template",
    status_code=status.HTTP_200_OK,
    summary="Descargar plantilla de movimientos de stock",
)
def bulk_movements_template():
    return template_response(
        "plantilla_movimientos_inventario.xlsx",
        BULK_MOVEMENT_HEADERS,
        BULK_MOVEMENT_EXAMPLES,
    )
