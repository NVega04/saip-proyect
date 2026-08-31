import csv
import io
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional

from fastapi import HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.exc import IntegrityError
from sqlmodel import Session

from src.schemas.schemas import BulkImportResult, BulkRowError

MAX_BULK_ROWS = 2000
MAX_FILE_MB = 5

_TRANSLIT = str.maketrans("áéíóúñÁÉÍÓÚÑ", "aeiounAEIOUN")


class RowError(Exception):
    def __init__(self, message: str):
        super().__init__(message)
        self.message = message


def _clean_db_error(exc: IntegrityError) -> str:
    orig = getattr(exc, "orig", None)
    if orig is not None:
        return str(orig)[:200]
    return str(exc)[:200]


def normalize_header(value: Any) -> str:
    raw = str(value or "").strip().translate(_TRANSLIT)
    return raw.lower().replace(" ", "_").replace("-", "_")


def name_key(value: Any) -> str:
    return str(value or "").strip().upper()


def safe_float(value: Any, label: str, default: float = 0) -> float:
    if value is None or str(value).strip() == "":
        return default
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        raise RowError(f"El campo '{label}' debe ser numérico.")


def required(row: Dict[str, Any], key: str, label: str) -> str:
    value = name_key(row.get(key))
    if not value:
        raise RowError(f"Falta el campo obligatorio '{label}'.")
    return value


def required_raw(row: Dict[str, Any], key: str, label: str) -> str:
    value = str(row.get(key) or "").strip()
    if not value:
        raise RowError(f"Falta el campo obligatorio '{label}'.")
    return value


def optional_text(row: Dict[str, Any], key: str) -> Optional[str]:
    value = row.get(key)
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip()
    if len(text) > 500:
        raise RowError(f"El campo '{key}' supera los 500 caracteres.")
    return text


def parse_datetime(value: Any, label: str) -> Optional[datetime]:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    formats = (
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M:%S",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise RowError(f"El campo '{label}' tiene una fecha inválida: '{text}'")


def _parse_rows(raw_rows) -> List[Dict[str, Any]]:
    headers: List[str] = []
    out: List[Dict[str, Any]] = []
    for values in raw_rows:
        cells = list(values) if not isinstance(values, (list, tuple)) else list(values)
        if all(c is None or str(c).strip() == "" for c in cells):
            continue
        if not headers:
            headers = [normalize_header(c) for c in cells]
            continue
        row: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            if idx < len(cells) and cells[idx] is not None:
                row[header] = cells[idx]
        out.append(row)
    return out


def parse_spreadsheet(file: UploadFile) -> List[Dict[str, Any]]:
    raw = file.file.read()
    if len(raw) > MAX_FILE_MB * 1024 * 1024:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"El archivo supera el límite de {MAX_FILE_MB} MB.",
        )

    filename = (file.filename or "").lower()
    try:
        if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            rows = _parse_rows(wb.active.iter_rows(values_only=True))
        elif filename.endswith(".csv"):
            text = raw.decode("utf-8-sig", errors="replace")
            rows = _parse_rows(csv.reader(io.StringIO(text)))
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Solo se aceptan archivos con extensión .xlsx o .csv.",
            )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"No se pudo leer el archivo: {exc}",
        )

    if len(rows) > MAX_BULK_ROWS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"El archivo supera el límite de {MAX_BULK_ROWS} filas de datos.",
        )
    return rows


def run_bulk(
    session: Session,
    rows: List[Dict[str, Any]],
    process_row: Callable[[Dict[str, Any]], Optional[bool]],
    file_line: Optional[Callable[[Dict[str, Any]], Optional[int]]] = None,
) -> BulkImportResult:
    creados = 0
    actualizados = 0
    errores: List[BulkRowError] = []
    for idx, row in enumerate(rows, start=2):
        try:
            with session.begin_nested():
                was_update = process_row(row)
                session.flush()
            if was_update:
                actualizados += 1
            else:
                creados += 1
        except RowError as exc:
            errores.append(BulkRowError(fila=file_line(row) if file_line else idx, mensaje=exc.message))
        except IntegrityError as exc:
            errores.append(
                BulkRowError(
                    fila=file_line(row) if file_line else idx,
                    mensaje=f"Violación de unicidad o integridad: {_clean_db_error(exc)}",
                )
            )
        except Exception as exc:
            errores.append(
                BulkRowError(fila=file_line(row) if file_line else idx, mensaje=str(exc))
            )
    session.commit()
    return BulkImportResult(
        total=len(rows), creados=creados, actualizados=actualizados, errores=errores
    )


def build_template_bytes(headers: List[str], example_rows: List[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in example_rows:
        ws.append(row)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 24
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()


def template_response(
    filename: str, headers: List[str], example_rows: List[list]
) -> StreamingResponse:
    data = build_template_bytes(headers, example_rows)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )