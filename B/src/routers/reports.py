from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from io import BytesIO
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.database import get_session
from src.dependencies import get_current_user
from src.models.models import User, Role

router = APIRouter(prefix="/reports", tags=["Reports"])

# ── Paleta SAIP ────────────────────────────────────────────────────────────
COLOR_HEADER_BG  = "5C3D1E"
COLOR_HEADER_FG  = "FFFFFF"
COLOR_ROW_ALT    = "F5F0EA"
COLOR_ROW_WHITE  = "FFFFFF"
COLOR_BORDER     = "CFC0B0"

ENTITY_MAP = {
    "users":    "Usuarios",
    "roles":    "Roles",
    "sessions": "Sesiones",
}

# ── Helpers de estilo ──────────────────────────────────────────────────────
def _header_font():
    return Font(name="Arial", bold=True, color=COLOR_HEADER_FG, size=10)

def _cell_font():
    return Font(name="Arial", size=9)

def _header_fill():
    return PatternFill("solid", fgColor=COLOR_HEADER_BG)

def _alt_fill():
    return PatternFill("solid", fgColor=COLOR_ROW_ALT)

def _white_fill():
    return PatternFill("solid", fgColor=COLOR_ROW_WHITE)

def _border():
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)

def _center():
    return Alignment(horizontal="center", vertical="center")

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _apply_header_row(ws, headers: list[str]):
    ws.row_dimensions[1].height = 22
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font   = _header_font()
        cell.fill   = _header_fill()
        cell.border = _border()
        cell.alignment = _center()


def _apply_data_row(ws, row_idx: int, values: list):
    fill = _alt_fill() if row_idx % 2 == 0 else _white_fill()
    ws.row_dimensions[row_idx].height = 18
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font      = _cell_font()
        cell.fill      = fill
        cell.border    = _border()
        cell.alignment = _left()


def _auto_width(ws, headers: list[str]):
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header) + 4
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[col_letter].width = min(max_len, 40)


def _add_title_row(ws, title: str, col_count: int):
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font      = Font(name="Arial", bold=True, size=12, color=COLOR_HEADER_FG)
    cell.fill      = _header_fill()
    cell.alignment = _center()
    ws.row_dimensions[1].height = 28


def _fmt_dt(value) -> str:
    if value is None:
        return "—"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value)


def _fmt_bool(value: bool) -> str:
    return "Sí" if value else "No"


# ── Builders por entidad ───────────────────────────────────────────────────
def _build_users(ws, db: Session):
    headers = ["ID", "Nombre", "Apellido", "Correo", "Teléfono", "Rol",
               "Admin", "Estado", "Creado en"]
    _apply_header_row(ws, headers)

    users = db.exec(select(User).where(User.deleted_at == None)).all()
    for i, u in enumerate(users, start=2):
        role_name = u.role.name if u.role else "—"
        _apply_data_row(ws, i, [
            u.id, u.first_name, u.last_name, u.email,
            u.phone or "—", role_name,
            _fmt_bool(u.is_admin), u.status.value,
            _fmt_dt(u.created_at),
        ])
    return headers


def _build_roles(ws, db: Session):
    headers = ["ID", "Nombre", "Descripción", "Estado", "Creado en"]
    _apply_header_row(ws, headers)

    roles = db.exec(select(Role).where(Role.deleted_at == None)).all()
    for i, r in enumerate(roles, start=2):
        _apply_data_row(ws, i, [
            r.id, r.name, r.description,
            r.status.value, _fmt_dt(r.created_at),
        ])
    return headers

BUILDERS = {
    "users":    _build_users,
    "roles":    _build_roles,
}


# ── Endpoint ───────────────────────────────────────────────────────────────
@router.get(
    "/{entity}",
    summary="Generar reporte Excel por entidad",
    response_class=StreamingResponse,
)
def generate_report(
    entity: str,
    db: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    if entity not in BUILDERS:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Entidad '{entity}' no soportada. Opciones: {list(BUILDERS.keys())}",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = ENTITY_MAP.get(entity, entity.capitalize())
    ws.freeze_panes = "A3"

    headers = BUILDERS[entity](ws, db)

    col_count = len(headers)
    _add_title_row(
        ws,
        f"Reporte de {ENTITY_MAP.get(entity, entity)} — {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC",
        col_count,
    )
    _auto_width(ws, headers)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"reporte_{entity}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )